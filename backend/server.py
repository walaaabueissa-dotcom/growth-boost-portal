from dotenv import load_dotenv
from pathlib import Path
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

import os
import uuid
import logging
from datetime import datetime, timezone, timedelta
from typing import List, Optional

import bcrypt
import jwt
from fastapi import FastAPI, APIRouter, HTTPException, Depends, Request, Response, UploadFile, File, Form
from fastapi.responses import FileResponse
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, EmailStr

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

UPLOAD_DIR = ROOT_DIR / 'uploads'
UPLOAD_DIR.mkdir(exist_ok=True)

app = FastAPI(title="Boost Growth Portal API")
api = APIRouter(prefix="/api")

JWT_ALGORITHM = "HS256"
JWT_SECRET = os.environ["JWT_SECRET"]

def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()

def hash_password(p: str) -> str:
    return bcrypt.hashpw(p.encode(), bcrypt.gensalt()).decode()

def verify_password(p: str, h: str) -> bool:
    return bcrypt.checkpw(p.encode(), h.encode())

def create_token(data: dict, hours: int = 24) -> str:
    payload = {**data, "exp": datetime.now(timezone.utc) + timedelta(hours=hours)}
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def decode_token(token: str) -> dict:
    return jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])

async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth = request.headers.get("Authorization", "")
        if auth.startswith("Bearer "):
            token = auth[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = decode_token(token)
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")
    if payload.get("role") == "admin":
        user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password_hash": 0})
    else:
        user = await db.therapists.find_one({"id": payload["sub"]}, {"_id": 0, "pin_hash": 0})
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    user["role"] = payload.get("role", user.get("role", "therapist"))
    return user

async def admin_only(user: dict = Depends(get_current_user)) -> dict:
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    return user

def set_auth_cookie(response: Response, token: str):
    response.set_cookie(key="access_token", value=token, httponly=True,
                        secure=True, samesite="none", max_age=86400, path="/")

# ------------------- Models -------------------
class LoginIn(BaseModel):
    email: EmailStr
    password: str

class TherapistPinLogin(BaseModel):
    therapist_id: str
    pin: str

class TherapistIn(BaseModel):
    name: str
    email: Optional[str] = None
    phone: Optional[str] = None
    color: Optional[str] = "#7A8A6A"
    pin: str

class TherapistUpdate(BaseModel):
    name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    color: Optional[str] = None
    pin: Optional[str] = None

class ScheduleCellIn(BaseModel):
    therapist_id: str
    day: int
    time_slot: str
    service_code: Optional[str] = "SS"
    child_name: Optional[str] = None
    note: Optional[str] = None
    custom_time: Optional[str] = None
    state: Optional[str] = "normal"
    color: Optional[str] = None
    duration: Optional[int] = 1  # number of time-slot rows the cell spans (1=single)
    week_start: str

class LocationIn(BaseModel):
    service: str
    address: str

class ClientIn(BaseModel):
    name: str
    file_no: Optional[str] = None
    age: Optional[str] = None
    parent_name: Optional[str] = None
    parent_phone: Optional[str] = None
    package_hours: Optional[float] = 24
    notes: Optional[str] = None
    main_therapist_id: Optional[str] = None
    co_therapist_ids: Optional[List[str]] = []
    supervisor: Optional[str] = None
    locations: Optional[List[LocationIn]] = []
    color: Optional[str] = None
    drive_url: Optional[str] = None

class SessionIn(BaseModel):
    client_id: str
    session_date: str  # ISO date
    start_time: Optional[str] = None  # "14:00"
    end_time: Optional[str] = None
    hours: float = 0
    status: str = "Completed"  # Completed, No Service, Cancelled, No Show
    therapist_ids: List[str] = []
    note: Optional[str] = None
    location: Optional[str] = None  # which location used (HS / SS)

class RequestIn(BaseModel):
    title: str
    description: Optional[str] = ""
    request_type: str = "general"
    priority: str = "normal"
    date_from: Optional[str] = None
    date_to: Optional[str] = None
    reward_type: Optional[str] = None
    extra_notes: Optional[str] = None

class RequestStatusUpdate(BaseModel):
    status: str
    admin_note: Optional[str] = None

class DirectoryContactIn(BaseModel):
    name: str
    role: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None

class IntakeIn(BaseModel):
    child_name: str
    parent_name: Optional[str] = None
    phone: Optional[str] = None
    intake_type: str = "pre"
    notes: Optional[str] = None
    status: Optional[str] = "new"
    intake_date: Optional[str] = None
    age: Optional[str] = None
    service: Optional[str] = None          # HS / SS / HS / SS
    district: Optional[str] = None          # Dis column
    time_pref: Optional[str] = None         # Morning / Evening / Any
    diagnosis: Optional[str] = None
    language: Optional[str] = None          # Post-intake only
    priority: Optional[bool] = False

class ResourceIn(BaseModel):
    title: str
    description: Optional[str] = None
    url: str
    category: Optional[str] = "drive"       # drive / file / link
    visibility: str = "all"                 # all / admin / therapist
    icon: Optional[str] = "Folders"
    bg: Optional[str] = "#E5EBE1"
    color: Optional[str] = "#3D4F35"
    sort_order: Optional[int] = 100

class DirectoryContactUpdate(BaseModel):
    name: Optional[str] = None
    role: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None

class LeaveIn(BaseModel):
    therapist_id: str
    start_date: str               # ISO yyyy-mm-dd
    end_date: str
    days: float = 1
    leave_type: Optional[str] = "Annual"   # Annual / Unpaid / Sickleave / Exam / Emergency
    status: Optional[str] = "pending"      # pending / approved / done / rejected / cancelled
    notes: Optional[str] = None
    admin_note: Optional[str] = None

class LeaveStatusUpdate(BaseModel):
    status: str
    admin_note: Optional[str] = None

class CancelNotifyIn(BaseModel):
    cell_id: str
    state: str                     # cancel_therapist / cancel_child
    message: str
    send_email: Optional[bool] = False
    extra_email: Optional[str] = None     # override or extra recipient

# ------------------- Auth -------------------
@api.post("/auth/login")
async def admin_login(payload: LoginIn, response: Response):
    email = payload.email.lower()
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(payload.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    token = create_token({"sub": user["id"], "role": "admin", "email": email})
    set_auth_cookie(response, token)
    return {"id": user["id"], "email": email, "name": user.get("name"), "role": "admin", "token": token}

@api.get("/auth/therapists-list")
async def therapists_list_public():
    return await db.therapists.find({}, {"_id": 0, "pin_hash": 0}).sort("name", 1).to_list(500)

@api.post("/auth/therapist-login")
async def therapist_login(payload: TherapistPinLogin, response: Response):
    t = await db.therapists.find_one({"id": payload.therapist_id})
    if not t or not verify_password(payload.pin, t["pin_hash"]):
        raise HTTPException(status_code=401, detail="Incorrect PIN")
    token = create_token({"sub": t["id"], "role": "therapist", "name": t["name"]})
    set_auth_cookie(response, token)
    return {"id": t["id"], "name": t["name"], "color": t.get("color"), "role": "therapist", "token": token}

@api.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    return user

@api.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    return {"ok": True}

# ------------------- Therapists -------------------
@api.get("/therapists")
async def list_therapists(user=Depends(get_current_user)):
    return await db.therapists.find({}, {"_id": 0, "pin_hash": 0}).sort("name", 1).to_list(500)

@api.post("/therapists")
async def create_therapist(payload: TherapistIn, _=Depends(admin_only)):
    tid = str(uuid.uuid4())
    doc = {"id": tid, "name": payload.name, "email": payload.email, "phone": payload.phone,
           "color": payload.color or "#7A8A6A", "pin_hash": hash_password(payload.pin),
           "created_at": now_iso()}
    await db.therapists.insert_one(doc)
    return {k: v for k, v in doc.items() if k not in ("_id", "pin_hash")}

@api.put("/therapists/{tid}")
async def update_therapist(tid: str, payload: TherapistUpdate, _=Depends(admin_only)):
    update = {k: v for k, v in payload.model_dump().items() if v is not None and k != "pin"}
    if payload.pin:
        update["pin_hash"] = hash_password(payload.pin)
    if not update:
        raise HTTPException(status_code=400, detail="No fields")
    await db.therapists.update_one({"id": tid}, {"$set": update})
    return await db.therapists.find_one({"id": tid}, {"_id": 0, "pin_hash": 0})

@api.delete("/therapists/{tid}")
async def delete_therapist(tid: str, _=Depends(admin_only)):
    await db.therapists.delete_one({"id": tid})
    return {"ok": True}

# ------------------- Schedule -------------------
@api.get("/schedule")
async def list_schedule(week_start: Optional[str] = None, user=Depends(get_current_user)):
    q: dict = {}
    if week_start:
        q["week_start"] = week_start
    cells = await db.schedule_cells.find(q, {"_id": 0}).to_list(5000)
    return cells  # everyone sees full schedule (master view) per user request

async def _notify(user_id: str, ntype: str, title: str, message: str):
    await db.notifications.insert_one({
        "id": str(uuid.uuid4()), "user_id": user_id, "type": ntype,
        "title": title, "message": message, "read": False, "created_at": now_iso(),
    })

async def _notify_admins(ntype: str, title: str, message: str):
    """Send notification to all admin users."""
    admins = await db.users.find({"role": "admin"}, {"_id": 0, "id": 1}).to_list(50)
    for a in admins:
        await _notify(a["id"], ntype, title, message)

@api.post("/schedule")
async def create_schedule_cell(payload: ScheduleCellIn, _=Depends(admin_only)):
    cid = str(uuid.uuid4())
    doc = {"id": cid, **payload.model_dump(), "created_at": now_iso()}
    await db.schedule_cells.insert_one(doc)
    doc.pop("_id", None)
    if doc.get("therapist_id"):
        await _notify(doc["therapist_id"], "schedule", "New session added",
                      f"{doc.get('service_code')} | {doc.get('child_name') or ''} at {doc.get('time_slot')}")
    return doc

@api.put("/schedule/{cid}")
async def update_schedule_cell(cid: str, payload: ScheduleCellIn, _=Depends(admin_only)):
    update = payload.model_dump()
    await db.schedule_cells.update_one({"id": cid}, {"$set": update})
    cell = await db.schedule_cells.find_one({"id": cid}, {"_id": 0})
    if cell and cell.get("therapist_id"):
        title = "Schedule update"
        if cell.get("state") == "cancel_therapist":
            title = "Session marked as Therapist Cancellation"
            await _notify_admins("cancel_alert", "Therapist cancellation",
                                 f"{cell.get('child_name') or '—'} session on day {cell.get('day')} at {cell.get('time_slot')} marked Therapist Cancel")
        elif cell.get("state") == "cancel_child":
            title = "Session marked as Client Cancellation"
            await _notify_admins("cancel_alert", "Client cancellation",
                                 f"{cell.get('child_name') or '—'} session on day {cell.get('day')} at {cell.get('time_slot')} marked Client Cancel")
        await _notify(cell["therapist_id"], "schedule", title,
                      f"{cell.get('service_code')} | {cell.get('child_name') or ''} at {cell.get('time_slot')}")
    return cell

@api.post("/schedule/{cid}/duplicate")
async def duplicate_cell(cid: str, _=Depends(admin_only)):
    cell = await db.schedule_cells.find_one({"id": cid}, {"_id": 0})
    if not cell:
        raise HTTPException(status_code=404, detail="Not found")
    new_cell = {**cell, "id": str(uuid.uuid4()), "created_at": now_iso()}
    await db.schedule_cells.insert_one(new_cell)
    new_cell.pop("_id", None)
    return new_cell

@api.delete("/schedule/{cid}")
async def delete_schedule_cell(cid: str, _=Depends(admin_only)):
    await db.schedule_cells.delete_one({"id": cid})
    return {"ok": True}

@api.post("/schedule/{cid}/notify")
async def notify_schedule(cid: str, body: dict, _=Depends(admin_only)):
    cell = await db.schedule_cells.find_one({"id": cid}, {"_id": 0})
    if not cell or not cell.get("therapist_id"):
        raise HTTPException(status_code=400, detail="No therapist assigned")
    msg = body.get("message") or f"Notice about session: {cell.get('child_name') or ''}"
    await _notify(cell["therapist_id"], "schedule_alert", "Notice from Admin", msg)
    return {"ok": True}

# ------------------- Clients & Sessions -------------------
@api.get("/clients")
async def list_clients(user=Depends(get_current_user)):
    if user.get("role") == "admin":
        return await db.clients.find({}, {"_id": 0}).sort("file_no", 1).to_list(500)
    # therapist: see only assigned (main or co)
    items = await db.clients.find({}, {"_id": 0}).sort("file_no", 1).to_list(500)
    uid = user["id"]
    return [c for c in items if c.get("main_therapist_id") == uid or uid in (c.get("co_therapist_ids") or [])]

@api.post("/clients")
async def create_client(payload: ClientIn, _=Depends(admin_only)):
    cid = str(uuid.uuid4())
    data = payload.model_dump()
    data["locations"] = [l for l in (data.get("locations") or [])]
    doc = {"id": cid, **data, "created_at": now_iso()}
    await db.clients.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.put("/clients/{cid}")
async def update_client(cid: str, payload: ClientIn, _=Depends(admin_only)):
    data = payload.model_dump()
    data["locations"] = [l for l in (data.get("locations") or [])]
    await db.clients.update_one({"id": cid}, {"$set": data})
    return await db.clients.find_one({"id": cid}, {"_id": 0})

@api.delete("/clients/{cid}")
async def delete_client(cid: str, _=Depends(admin_only)):
    await db.clients.delete_one({"id": cid})
    await db.sessions.delete_many({"client_id": cid})
    return {"ok": True}

# ------------------- Sessions (Attendance log) -------------------
@api.get("/sessions")
async def list_sessions(client_id: Optional[str] = None, user=Depends(get_current_user)):
    q = {}
    if client_id:
        q["client_id"] = client_id
    items = await db.sessions.find(q, {"_id": 0}).sort("session_date", -1).to_list(2000)
    if user.get("role") == "therapist":
        uid = user["id"]
        items = [s for s in items if uid in (s.get("therapist_ids") or [])]
    return items

@api.post("/sessions")
async def create_session(payload: SessionIn, user=Depends(get_current_user)):
    sid = str(uuid.uuid4())
    therapist_ids = payload.therapist_ids or []
    if user.get("role") == "therapist" and user["id"] not in therapist_ids:
        therapist_ids.append(user["id"])
    doc = {"id": sid, **payload.model_dump(), "therapist_ids": therapist_ids,
           "created_by": user["id"], "created_by_role": user["role"],
           "created_at": now_iso()}
    await db.sessions.insert_one(doc)
    doc.pop("_id", None)
    # Admin alerts
    client = await db.clients.find_one({"id": payload.client_id}, {"_id": 0})
    cname = client.get("name") if client else "—"
    if user.get("role") == "therapist":
        await _notify_admins("session_log", f"New session logged ({payload.status})",
                             f"{user.get('name')} logged {payload.status} for {cname} ({payload.hours}h)")
    if payload.status in ("Cancelled", "No Show"):
        await _notify_admins("cancel_alert", f"Session {payload.status}: {cname}",
                             f"On {payload.session_date} ({user.get('name')})")
    # Low-hours alert
    if client:
        used = await db.sessions.aggregate([
            {"$match": {"client_id": payload.client_id, "status": "Completed"}},
            {"$group": {"_id": None, "total": {"$sum": "$hours"}}}
        ]).to_list(1)
        used_h = used[0]["total"] if used else 0
        rem = (client.get("package_hours") or 24) - used_h
        if 0 < rem <= 4:
            await _notify_admins("low_hours", f"⚠️ {cname} has only {rem}h left",
                                 f"Pkg {client.get('package_hours')}h, used {used_h}h. Consider package renewal.")
        elif rem <= 0:
            await _notify_admins("low_hours", f"🔴 {cname} package exhausted",
                                 f"Used {used_h}h of {client.get('package_hours')}h.")
    return doc

@api.put("/sessions/{sid}")
async def update_session(sid: str, payload: SessionIn, user=Depends(get_current_user)):
    sess = await db.sessions.find_one({"id": sid})
    if not sess:
        raise HTTPException(status_code=404, detail="Not found")
    if user.get("role") != "admin" and user["id"] not in (sess.get("therapist_ids") or []):
        raise HTTPException(status_code=403, detail="Forbidden")
    await db.sessions.update_one({"id": sid}, {"$set": payload.model_dump()})
    return await db.sessions.find_one({"id": sid}, {"_id": 0})

@api.delete("/sessions/{sid}")
async def delete_session(sid: str, user=Depends(get_current_user)):
    sess = await db.sessions.find_one({"id": sid})
    if not sess:
        return {"ok": True}
    if user.get("role") != "admin" and user["id"] not in (sess.get("therapist_ids") or []):
        raise HTTPException(status_code=403, detail="Forbidden")
    await db.sessions.delete_one({"id": sid})
    return {"ok": True}

# ------------------- Attendance Sheets (file upload, kept for backward compat) -------------------
@api.get("/clients/{cid}/sheets")
async def list_sheets(cid: str, user=Depends(get_current_user)):
    return await db.attendance_sheets.find({"client_id": cid}, {"_id": 0}).sort("page_number", 1).to_list(500)

@api.post("/clients/{cid}/sheets")
async def upload_sheet(cid: str,
                      title: str = Form(...),
                      session_date: str = Form(...),
                      therapist_id: Optional[str] = Form(None),
                      notes: Optional[str] = Form(None),
                      file: Optional[UploadFile] = File(None),
                      _=Depends(admin_only)):
    sid = str(uuid.uuid4())
    file_path = None
    file_name = None
    if file:
        ext = Path(file.filename).suffix
        file_name = file.filename
        save_path = UPLOAD_DIR / f"{sid}{ext}"
        save_path.write_bytes(await file.read())
        file_path = f"{sid}{ext}"
    last = await db.attendance_sheets.find_one({"client_id": cid}, sort=[("page_number", -1)])
    page_number = (last.get("page_number", 0) + 1) if last else 1
    doc = {"id": sid, "client_id": cid, "title": title, "session_date": session_date,
           "therapist_id": therapist_id, "notes": notes, "page_number": page_number,
           "file_name": file_name, "file_path": file_path, "created_at": now_iso()}
    await db.attendance_sheets.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.delete("/sheets/{sid}")
async def delete_sheet(sid: str, _=Depends(admin_only)):
    sheet = await db.attendance_sheets.find_one({"id": sid})
    if sheet and sheet.get("file_path"):
        fp = UPLOAD_DIR / sheet["file_path"]
        if fp.exists():
            fp.unlink()
    await db.attendance_sheets.delete_one({"id": sid})
    return {"ok": True}

@api.get("/sheets/{sid}/download")
async def download_sheet(sid: str, user=Depends(get_current_user)):
    sheet = await db.attendance_sheets.find_one({"id": sid}, {"_id": 0})
    if not sheet or not sheet.get("file_path"):
        raise HTTPException(status_code=404, detail="No file")
    fp = UPLOAD_DIR / sheet["file_path"]
    return FileResponse(str(fp), filename=sheet.get("file_name") or sheet["file_path"])

# ------------------- Requests -------------------
@api.get("/requests")
async def list_requests(user=Depends(get_current_user)):
    q = {} if user.get("role") == "admin" else {"therapist_id": user["id"]}
    return await db.requests.find(q, {"_id": 0}).sort("created_at", -1).to_list(500)

@api.post("/requests")
async def create_request(payload: RequestIn, user=Depends(get_current_user)):
    if user.get("role") != "therapist":
        raise HTTPException(status_code=403, detail="Therapist only")
    rid = str(uuid.uuid4())
    doc = {"id": rid, "therapist_id": user["id"], "therapist_name": user.get("name"),
           **payload.model_dump(), "status": "pending", "admin_note": None,
           "created_at": now_iso(), "updated_at": now_iso(),
           "timeline": [{"event": "submitted", "at": now_iso(), "by": user.get("name")}]}
    await db.requests.insert_one(doc)
    doc.pop("_id", None)
    # Notify admins of new request
    await _notify_admins("request_new", f"New {payload.request_type} request",
                         f"{user.get('name')}: {payload.title} (priority: {payload.priority})")
    return doc

@api.put("/requests/{rid}/status")
async def update_request_status(rid: str, payload: RequestStatusUpdate, admin=Depends(admin_only)):
    req = await db.requests.find_one({"id": rid})
    if not req:
        raise HTTPException(status_code=404, detail="Not found")
    timeline = req.get("timeline", [])
    timeline.append({"event": payload.status, "at": now_iso(), "by": admin.get("name") or "Admin",
                     "note": payload.admin_note})
    await db.requests.update_one({"id": rid}, {"$set": {
        "status": payload.status, "admin_note": payload.admin_note,
        "updated_at": now_iso(), "timeline": timeline,
    }})
    status_map = {"pending": "Pending", "in_progress": "In Progress",
                  "approved": "Approved", "rejected": "Rejected", "done": "Completed"}
    await _notify(req["therapist_id"], "request", "Request update",
                  f"Your request '{req['title']}' is now: {status_map.get(payload.status, payload.status)}")
    return await db.requests.find_one({"id": rid}, {"_id": 0})

@api.delete("/requests/{rid}")
async def delete_request(rid: str, user=Depends(get_current_user)):
    req = await db.requests.find_one({"id": rid})
    if not req:
        return {"ok": True}
    if user.get("role") != "admin" and req.get("therapist_id") != user["id"]:
        raise HTTPException(status_code=403, detail="Forbidden")
    await db.requests.delete_one({"id": rid})
    return {"ok": True}

# ------------------- Notifications -------------------
@api.get("/notifications")
async def list_notifications(user=Depends(get_current_user)):
    return await db.notifications.find({"user_id": user["id"]}, {"_id": 0}).sort("created_at", -1).limit(100).to_list(100)

@api.post("/notifications/{nid}/read")
async def mark_read(nid: str, user=Depends(get_current_user)):
    await db.notifications.update_one({"id": nid, "user_id": user["id"]}, {"$set": {"read": True}})
    return {"ok": True}

@api.post("/notifications/read-all")
async def mark_all_read(user=Depends(get_current_user)):
    await db.notifications.update_many({"user_id": user["id"]}, {"$set": {"read": True}})
    return {"ok": True}

# ------------------- Directory -------------------
@api.get("/directory")
async def list_directory(user=Depends(get_current_user)):
    return await db.directory.find({}, {"_id": 0}).to_list(500)

@api.post("/directory")
async def create_contact(payload: DirectoryContactIn, _=Depends(admin_only)):
    cid = str(uuid.uuid4())
    doc = {"id": cid, **payload.model_dump(), "created_at": now_iso()}
    await db.directory.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.put("/directory/{cid}")
async def update_contact(cid: str, payload: DirectoryContactUpdate, _=Depends(admin_only)):
    update = {k: v for k, v in payload.model_dump().items() if v is not None}
    if not update:
        raise HTTPException(status_code=400, detail="No fields")
    await db.directory.update_one({"id": cid}, {"$set": update})
    return await db.directory.find_one({"id": cid}, {"_id": 0})

@api.delete("/directory/{cid}")
async def delete_contact(cid: str, _=Depends(admin_only)):
    await db.directory.delete_one({"id": cid})
    return {"ok": True}

# ------------------- Resources -------------------
@api.get("/resources")
async def list_resources(user=Depends(get_current_user)):
    items = await db.resources.find({}, {"_id": 0}).sort("sort_order", 1).to_list(500)
    if user.get("role") == "admin":
        return items
    # Therapists see only "therapist" and "all" visibility
    return [r for r in items if r.get("visibility") in ("therapist", "all")]

@api.post("/resources")
async def create_resource(payload: ResourceIn, _=Depends(admin_only)):
    rid = str(uuid.uuid4())
    doc = {"id": rid, **payload.model_dump(), "created_at": now_iso()}
    await db.resources.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.put("/resources/{rid}")
async def update_resource(rid: str, payload: ResourceIn, _=Depends(admin_only)):
    await db.resources.update_one({"id": rid}, {"$set": payload.model_dump()})
    return await db.resources.find_one({"id": rid}, {"_id": 0})

@api.get("/clients/{cid}/sessions/export")
async def export_sessions_excel(cid: str, user=Depends(get_current_user)):
    """Export client's session history as Excel sheet (Boost Growth Attendance Sheet style)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO

    client = await db.clients.find_one({"id": cid}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    if user.get("role") == "therapist" and user["id"] not in (client.get("co_therapist_ids") or []) + ([client.get("main_therapist_id")] if client.get("main_therapist_id") else []):
        raise HTTPException(status_code=403, detail="Forbidden")

    sessions = await db.sessions.find({"client_id": cid}, {"_id": 0}).sort("session_date", -1).to_list(2000)
    therapists = {t["id"]: t for t in await db.therapists.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(100)}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"
    head_fill = PatternFill("solid", fgColor="7A8A6A")
    head_font = Font(bold=True, color="FFFFFF", size=11)
    sub_fill = PatternFill("solid", fgColor="EFE8D2")
    border = Border(left=Side(style="thin", color="B5B0A0"), right=Side(style="thin", color="B5B0A0"),
                    top=Side(style="thin", color="B5B0A0"), bottom=Side(style="thin", color="B5B0A0"))
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Title row
    ws.merge_cells("A1:G1")
    ws["A1"] = f"BOOST GROWTH — Attendance Sheet — {client.get('name')} (#{client.get('file_no') or '—'})"
    ws["A1"].font = Font(bold=True, size=14, color="2C3625")
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 28

    # Patient row
    used = sum(float(s.get("hours") or 0) for s in sessions if s.get("status") == "Completed")
    pkg = float(client.get("package_hours") or 24)
    rem = max(0.0, pkg - used)
    ws["A2"] = "Client"; ws["B2"] = client.get("name")
    ws["C2"] = "File #"; ws["D2"] = client.get("file_no") or "—"
    ws["E2"] = "Package"; ws["F2"] = f"{pkg}h"
    ws["G2"] = f"Used {used}h · Remaining {rem}h"
    for c in "ABCDEFG":
        ws[f"{c}2"].fill = sub_fill
        ws[f"{c}2"].font = Font(bold=True, color="2C3625")
        ws[f"{c}2"].alignment = center

    # Header
    headers = ["Date", "Day", "Status", "Time", "Hours", "Therapist", "Note"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=i, value=h)
        cell.fill = head_fill; cell.font = head_font; cell.alignment = center; cell.border = border
    ws.row_dimensions[4].height = 24

    # Session rows
    DAY_NAMES = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
    STATUS_FILLS = {"Completed": "D9EAD3", "Cancelled": "FCE0E8", "No Show": "FFF4C4", "No Service": "ECECEC"}
    row = 5
    total_completed = total_cancelled = total_no_show = total_hours = 0
    for s in sessions:
        sd = s.get("session_date") or ""
        try:
            day_label = DAY_NAMES[datetime.fromisoformat(sd).weekday()] if sd else "—"
        except Exception:
            day_label = "—"
        therapist_names = ", ".join((therapists.get(tid) or {}).get("name", "?") for tid in (s.get("therapist_ids") or []))
        ws.cell(row=row, column=1, value=sd).alignment = center
        ws.cell(row=row, column=2, value=day_label).alignment = center
        st_cell = ws.cell(row=row, column=3, value=s.get("status") or "—")
        st_cell.alignment = center
        if s.get("status") in STATUS_FILLS:
            st_cell.fill = PatternFill("solid", fgColor=STATUS_FILLS[s["status"]])
        time_str = ""
        if s.get("start_time") and s.get("end_time"):
            time_str = f"{s['start_time']} – {s['end_time']}"
        ws.cell(row=row, column=4, value=time_str).alignment = center
        ws.cell(row=row, column=5, value=float(s.get("hours") or 0)).alignment = center
        ws.cell(row=row, column=6, value=therapist_names).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.cell(row=row, column=7, value=s.get("note") or "").alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = border
        if s.get("status") == "Completed":
            total_completed += 1; total_hours += float(s.get("hours") or 0)
        elif s.get("status") == "Cancelled":
            total_cancelled += 1
        elif s.get("status") == "No Show":
            total_no_show += 1
        row += 1

    # Footer totals
    foot = row + 1
    ws.cell(row=foot, column=1, value="TOTALS").font = Font(bold=True, color="2C3625")
    ws.cell(row=foot, column=2, value=f"Delivered: {total_completed}")
    ws.cell(row=foot, column=3, value=f"Cancelled: {total_cancelled}")
    ws.cell(row=foot, column=4, value=f"No-Show: {total_no_show}")
    ws.cell(row=foot, column=5, value=total_hours).alignment = center
    ws.cell(row=foot, column=6, value=f"Hours Remaining: {round(rem, 1)}")
    for col in range(1, 8):
        ws.cell(row=foot, column=col).fill = sub_fill
        ws.cell(row=foot, column=col).font = Font(bold=True, color="2C3625")

    # Column widths
    widths = [12, 8, 14, 16, 8, 24, 32]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    from fastapi.responses import Response
    fname = f"attendance_{client.get('file_no') or 'client'}_{client.get('name','').replace(' ','_')}.xlsx"
    return Response(content=out.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename={fname}"})

# ------------------- Email Settings (admin) -------------------
class EmailSettingsIn(BaseModel):
    resend_api_key: Optional[str] = None
    from_email: Optional[str] = None

@api.get("/admin/email-settings")
async def get_email_settings(_=Depends(admin_only)):
    doc = await db.settings.find_one({"key": "email"}, {"_id": 0}) or {}
    has_key = bool(doc.get("resend_api_key") or os.environ.get("RESEND_API_KEY"))
    return {
        "configured": has_key,
        "from_email": doc.get("from_email") or os.environ.get("EMAIL_FROM") or "Boost Growth <noreply@boostgrowthsa.com>",
        "key_preview": (doc.get("resend_api_key") or "")[:8] + "..." if doc.get("resend_api_key") else None,
    }

@api.post("/admin/email-settings")
async def save_email_settings(payload: EmailSettingsIn, _=Depends(admin_only)):
    update = {k: v for k, v in payload.model_dump().items() if v}
    if not update:
        raise HTTPException(status_code=400, detail="No fields")
    update["updated_at"] = now_iso()
    await db.settings.update_one({"key": "email"}, {"$set": update, "$setOnInsert": {"key": "email"}}, upsert=True)
    # Also set into env so existing _send_email_stub picks it up
    if "resend_api_key" in update:
        os.environ["RESEND_API_KEY"] = update["resend_api_key"]
    if "from_email" in update:
        os.environ["EMAIL_FROM"] = update["from_email"]
    return {"ok": True, "configured": True}

@api.get("/admin/email-queue")
async def list_email_queue(_=Depends(admin_only)):
    return await db.email_queue.find({}, {"_id": 0}).sort("created_at", -1).limit(100).to_list(100)

@api.delete("/resources/{rid}")
async def delete_resource(rid: str, _=Depends(admin_only)):
    await db.resources.delete_one({"id": rid})
    return {"ok": True}

# ------------------- Leaves / Vacations -------------------
DEFAULT_ANNUAL_BALANCE = 30  # baseline annual leave per year

@api.get("/leaves")
async def list_leaves(year: Optional[int] = None, user=Depends(get_current_user)):
    q: dict = {}
    if year:
        q["start_date"] = {"$gte": f"{year}-01-01", "$lte": f"{year}-12-31"}
    if user.get("role") != "admin":
        q["therapist_id"] = user["id"]
    items = await db.leaves.find(q, {"_id": 0}).sort("start_date", -1).to_list(2000)
    # Enrich with therapist name + email for admin
    if user.get("role") == "admin":
        therapists = await db.therapists.find({}, {"_id": 0, "id": 1, "name": 1, "email": 1, "color": 1}).to_list(100)
        t_by_id = {t["id"]: t for t in therapists}
        for it in items:
            t = t_by_id.get(it.get("therapist_id"))
            if t:
                it["therapist_name"] = t.get("name")
                it["therapist_color"] = t.get("color")
                it["therapist_email"] = t.get("email")
    return items

@api.get("/leaves/balance")
async def leaves_balance(year: Optional[int] = None, user=Depends(get_current_user)):
    """Return per-therapist annual balance: {therapist_id, name, allocated, used (Annual+approved/done), remaining, breakdown}.
    For therapist role: only their own.
    """
    yr = year or datetime.now(timezone.utc).year
    therapists = await db.therapists.find({}, {"_id": 0, "id": 1, "name": 1, "color": 1, "email": 1, "annual_balance": 1}).to_list(100)
    if user.get("role") != "admin":
        therapists = [t for t in therapists if t["id"] == user["id"]]
    leaves = await db.leaves.find({"start_date": {"$gte": f"{yr}-01-01", "$lte": f"{yr}-12-31"}}, {"_id": 0}).to_list(2000)
    out = []
    for t in therapists:
        own = [l for l in leaves if l.get("therapist_id") == t["id"]]
        used_annual = sum(float(l.get("days") or 0) for l in own if l.get("leave_type") == "Annual" and l.get("status") in ("approved", "done"))
        used_unpaid = sum(float(l.get("days") or 0) for l in own if l.get("leave_type") == "Unpaid" and l.get("status") in ("approved", "done"))
        used_sick = sum(float(l.get("days") or 0) for l in own if l.get("leave_type") == "Sickleave" and l.get("status") in ("approved", "done"))
        pending = sum(float(l.get("days") or 0) for l in own if l.get("status") == "pending")
        allocated = float(t.get("annual_balance") or DEFAULT_ANNUAL_BALANCE)
        remaining = max(0.0, allocated - used_annual)
        out.append({
            "therapist_id": t["id"], "name": t["name"], "color": t.get("color"), "email": t.get("email"),
            "year": yr, "allocated": allocated,
            "used_annual": round(used_annual, 1),
            "used_unpaid": round(used_unpaid, 1),
            "used_sick": round(used_sick, 1),
            "pending": round(pending, 1),
            "remaining": round(remaining, 1),
            "leaves_count": len(own),
        })
    return out

@api.post("/leaves")
async def create_leave(payload: LeaveIn, user=Depends(get_current_user)):
    if user.get("role") != "admin" and payload.therapist_id != user["id"]:
        raise HTTPException(status_code=403, detail="Therapist can only create own leaves")
    lid = str(uuid.uuid4())
    doc = {"id": lid, **payload.model_dump(), "created_by": user["id"], "created_at": now_iso()}
    if user.get("role") != "admin":
        doc["status"] = "pending"  # therapist requests start as pending
    await db.leaves.insert_one(doc)
    doc.pop("_id", None)
    # Notify admins if therapist submitted
    if user.get("role") != "admin":
        await _notify_admins("leave_request", "New leave request",
                             f"{user.get('name')}: {payload.leave_type} {payload.days}d ({payload.start_date} → {payload.end_date})")
    return doc

@api.put("/leaves/{lid}")
async def update_leave(lid: str, payload: LeaveIn, user=Depends(get_current_user)):
    leave = await db.leaves.find_one({"id": lid})
    if not leave:
        raise HTTPException(status_code=404, detail="Not found")
    if user.get("role") != "admin" and leave.get("therapist_id") != user["id"]:
        raise HTTPException(status_code=403, detail="Forbidden")
    update = payload.model_dump()
    await db.leaves.update_one({"id": lid}, {"$set": update})
    return await db.leaves.find_one({"id": lid}, {"_id": 0})

@api.put("/leaves/{lid}/status")
async def update_leave_status(lid: str, payload: LeaveStatusUpdate, admin=Depends(admin_only)):
    leave = await db.leaves.find_one({"id": lid})
    if not leave:
        raise HTTPException(status_code=404, detail="Not found")
    await db.leaves.update_one({"id": lid}, {"$set": {
        "status": payload.status, "admin_note": payload.admin_note,
        "decided_by": admin.get("name") or "Admin", "decided_at": now_iso(),
    }})
    # Notify therapist
    if leave.get("therapist_id"):
        msg_map = {"approved": "Approved", "rejected": "Rejected", "done": "Completed", "cancelled": "Cancelled", "pending": "Pending"}
        await _notify(leave["therapist_id"], "leave",
                      f"Leave {msg_map.get(payload.status, payload.status)}",
                      f"Your {leave.get('leave_type')} leave from {leave.get('start_date')} to {leave.get('end_date')} ({leave.get('days')}d) is now {msg_map.get(payload.status, payload.status)}.")
    return await db.leaves.find_one({"id": lid}, {"_id": 0})

@api.delete("/leaves/{lid}")
async def delete_leave(lid: str, user=Depends(get_current_user)):
    leave = await db.leaves.find_one({"id": lid})
    if not leave:
        return {"ok": True}
    if user.get("role") != "admin" and leave.get("therapist_id") != user["id"]:
        raise HTTPException(status_code=403, detail="Forbidden")
    await db.leaves.delete_one({"id": lid})
    return {"ok": True}

# ------------------- Cancel-Notify (in-app + queued email) -------------------
async def _send_email_stub(to: str, subject: str, body: str) -> dict:
    """Email send stub. Will integrate with Resend once API key is configured.
    Currently logs and stores in db.email_queue for later delivery.
    """
    api_key = os.environ.get("RESEND_API_KEY")
    queue_doc = {
        "id": str(uuid.uuid4()),
        "to": to, "subject": subject, "body": body,
        "status": "queued", "provider": "resend",
        "created_at": now_iso(),
    }
    if api_key:
        try:
            import httpx
            async with httpx.AsyncClient() as cli:
                r = await cli.post("https://api.resend.com/emails",
                                   headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
                                   json={"from": os.environ.get("EMAIL_FROM", "Boost Growth <noreply@boostgrowthsa.com>"),
                                         "to": [to], "subject": subject, "html": f"<p>{body.replace(chr(10),'<br/>')}</p>"})
                if r.status_code in (200, 202):
                    queue_doc["status"] = "sent"
                    queue_doc["provider_id"] = r.json().get("id")
                else:
                    queue_doc["status"] = "failed"
                    queue_doc["error"] = r.text[:300]
        except Exception as e:
            queue_doc["status"] = "failed"
            queue_doc["error"] = str(e)[:300]
    else:
        queue_doc["status"] = "queued_no_key"
    await db.email_queue.insert_one(queue_doc)
    queue_doc.pop("_id", None)
    return queue_doc

@api.post("/schedule/cancel-notify")
async def schedule_cancel_notify(payload: CancelNotifyIn, _=Depends(admin_only)):
    """Mark cell as cancelled + send in-app notification + queue email if requested."""
    cell = await db.schedule_cells.find_one({"id": payload.cell_id}, {"_id": 0})
    if not cell:
        raise HTTPException(status_code=404, detail="Schedule cell not found")
    # Update state
    await db.schedule_cells.update_one({"id": payload.cell_id}, {"$set": {"state": payload.state}})
    # In-app notification
    if cell.get("therapist_id"):
        title = "Session marked as Therapist Cancellation" if payload.state == "cancel_therapist" else "Session marked as Client Cancellation"
        await _notify(cell["therapist_id"], "schedule_cancel", title, payload.message)
        # Email
        email_result = None
        if payload.send_email:
            therapist = await db.therapists.find_one({"id": cell["therapist_id"]}, {"_id": 0})
            recipient = payload.extra_email or (therapist.get("email") if therapist else None)
            if recipient:
                subj = f"[Boost Growth] {title}"
                body_lines = [
                    f"Hello {therapist.get('name') if therapist else ''},",
                    "",
                    payload.message,
                    "",
                    f"Cell: {cell.get('service_code')} | {cell.get('child_name') or '—'}",
                    f"Day: {cell.get('day')} | Time: {cell.get('time_slot')}",
                    "",
                    "— Boost Growth Portal",
                ]
                email_result = await _send_email_stub(recipient, subj, "\n".join(body_lines))
        return {"ok": True, "in_app": True, "email": email_result}
    return {"ok": True, "in_app": False}

# ------------------- Intake (admin only) -------------------
@api.get("/intake")
async def list_intake(_=Depends(admin_only)):
    return await db.intake.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)

@api.post("/intake")
async def create_intake(payload: IntakeIn, _=Depends(admin_only)):
    iid = str(uuid.uuid4())
    doc = {"id": iid, **payload.model_dump(), "created_at": now_iso()}
    await db.intake.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.put("/intake/{iid}")
async def update_intake(iid: str, payload: IntakeIn, _=Depends(admin_only)):
    await db.intake.update_one({"id": iid}, {"$set": payload.model_dump()})
    return await db.intake.find_one({"id": iid}, {"_id": 0})

@api.delete("/intake/{iid}")
async def delete_intake(iid: str, _=Depends(admin_only)):
    await db.intake.delete_one({"id": iid})
    return {"ok": True}

# ------------------- Reports -------------------
@api.get("/reports/dashboard")
async def reports_dashboard(_=Depends(admin_only)):
    sessions = await db.sessions.find({}, {"_id": 0}).to_list(5000)
    clients = await db.clients.find({}, {"_id": 0}).to_list(500)
    therapists = await db.therapists.find({}, {"_id": 0, "pin_hash": 0}).to_list(50)
    requests = await db.requests.find({}, {"_id": 0}).to_list(500)
    cells = await db.schedule_cells.find({}, {"_id": 0}).to_list(5000)

    # Sessions per therapist
    per_t: dict = {}
    for t in therapists:
        per_t[t["id"]] = {"name": t["name"], "color": t.get("color"),
                           "completed": 0, "cancelled": 0, "no_show": 0, "no_service": 0,
                           "hours": 0.0}
    for s in sessions:
        for tid in s.get("therapist_ids") or []:
            if tid in per_t:
                if s["status"] == "Completed":
                    per_t[tid]["completed"] += 1
                    per_t[tid]["hours"] += float(s.get("hours") or 0)
                elif s["status"] == "Cancelled":
                    per_t[tid]["cancelled"] += 1
                elif s["status"] == "No Show":
                    per_t[tid]["no_show"] += 1
                else:
                    per_t[tid]["no_service"] += 1

    # Per-client used hours + status
    per_c = []
    for c in clients:
        used = sum(float(s.get("hours") or 0) for s in sessions if s.get("client_id") == c["id"] and s.get("status") == "Completed")
        pkg = c.get("package_hours") or 24
        rem = max(0, pkg - used)
        if rem <= 0 or rem <= 2 or rem / pkg <= 0.2:
            status = "urgent"
        elif rem / pkg <= 0.35 or rem <= 4:
            status = "warning"
        else:
            status = "ok"
        per_c.append({"id": c["id"], "name": c["name"], "file_no": c.get("file_no"),
                      "color": c.get("color"), "pkg": pkg, "used": round(used, 1),
                      "rem": round(rem, 1), "status": status})

    # Cancellation breakdown from schedule cells (this week)
    sched_cancel_t = sum(1 for c in cells if c.get("state") == "cancel_therapist")
    sched_cancel_c = sum(1 for c in cells if c.get("state") == "cancel_child")

    return {
        "totals": {
            "therapists": len(therapists),
            "clients": len(clients),
            "sessions": len(sessions),
            "completed_sessions": sum(1 for s in sessions if s.get("status") == "Completed"),
            "total_hours": round(sum(float(s.get("hours") or 0) for s in sessions if s.get("status") == "Completed"), 1),
            "open_requests": sum(1 for r in requests if r.get("status") == "pending"),
            "urgent_clients": sum(1 for c in per_c if c["status"] == "urgent"),
            "warning_clients": sum(1 for c in per_c if c["status"] == "warning"),
            "schedule_cells": len(cells),
            "schedule_cancel_therapist": sched_cancel_t,
            "schedule_cancel_child": sched_cancel_c,
        },
        "per_therapist": list(per_t.values()),
        "per_client": sorted(per_c, key=lambda x: {"urgent":0,"warning":1,"ok":2}[x["status"]]),
    }

# ------------------- Imports -------------------
def _read_table(file: UploadFile) -> List[dict]:
    """Read xlsx/csv into list of dicts with normalized lower-case keys."""
    import pandas as pd
    content = file.file.read()
    import io
    if file.filename.lower().endswith(".csv"):
        df = pd.read_csv(io.BytesIO(content))
    else:
        df = pd.read_excel(io.BytesIO(content), engine="openpyxl")
    df.columns = [str(c).strip().lower() for c in df.columns]
    df = df.where(df.notna(), None)
    return df.to_dict("records")

@api.post("/import/clients")
async def import_clients(file: UploadFile = File(...), _=Depends(admin_only)):
    rows = _read_table(file)
    created, skipped = 0, 0
    therapists = await db.therapists.find({}, {"_id": 0}).to_list(100)
    t_by_name = {t["name"].lower(): t["id"] for t in therapists}
    for r in rows:
        name = r.get("name") or r.get("child_name") or r.get("full_name")
        if not name:
            skipped += 1; continue
        file_no = str(r.get("file_no") or r.get("id") or r.get("file") or "").strip() or None
        # match therapist name to id
        main_name = (r.get("main_therapist") or r.get("main") or "").strip().lower() if r.get("main_therapist") or r.get("main") else None
        main_id = t_by_name.get(main_name) if main_name else None
        await db.clients.insert_one({
            "id": str(uuid.uuid4()), "name": str(name).strip(),
            "file_no": file_no, "package_hours": float(r.get("package_hours") or r.get("pkg") or 24),
            "supervisor": r.get("supervisor"), "main_therapist_id": main_id,
            "co_therapist_ids": [], "color": r.get("color") or "#A2C4C9",
            "locations": [], "parent_name": r.get("parent_name") or r.get("parent"),
            "parent_phone": str(r.get("parent_phone") or r.get("phone") or "") or None,
            "age": str(r.get("age") or "") or None, "notes": r.get("notes"),
            "created_at": now_iso(),
        })
        created += 1
    return {"created": created, "skipped": skipped}

@api.post("/import/intake")
async def import_intake(file: UploadFile = File(...), _=Depends(admin_only)):
    rows = _read_table(file)
    created, skipped = 0, 0
    for r in rows:
        name = r.get("child_name") or r.get("name")
        if not name:
            skipped += 1; continue
        intake_type = (r.get("intake_type") or r.get("type") or "pre").lower()
        if intake_type not in ("pre", "post"):
            intake_type = "pre"
        await db.intake.insert_one({
            "id": str(uuid.uuid4()), "child_name": str(name).strip(),
            "parent_name": r.get("parent_name") or r.get("parent"),
            "phone": str(r.get("phone") or "") or None,
            "intake_type": intake_type, "status": (r.get("status") or "new").lower(),
            "notes": r.get("notes"), "intake_date": str(r.get("intake_date") or "") or None,
            "age": str(r.get("age") or "") or None, "created_at": now_iso(),
        })
        created += 1
    return {"created": created, "skipped": skipped}

# ------------------- Historical Schedule Loader -------------------
HISTORICAL_SCHEDULES = None  # lazy-loaded from JSON file

def _load_historical():
    global HISTORICAL_SCHEDULES
    if HISTORICAL_SCHEDULES is None:
        import json
        path = ROOT_DIR / "historical_schedules.json"
        if path.exists():
            HISTORICAL_SCHEDULES = json.loads(path.read_text())
        else:
            HISTORICAL_SCHEDULES = {}
    return HISTORICAL_SCHEDULES

@api.get("/import/historical-weeks")
async def list_historical_weeks(_=Depends(admin_only)):
    data = _load_historical()
    return {"weeks": list(data.keys())}

@api.post("/import/historical-load")
async def import_historical(body: dict, _=Depends(admin_only)):
    """Import all historical weeks into schedule_cells. body: {clear_existing?: bool}"""
    data = _load_historical()
    if not data:
        raise HTTPException(status_code=404, detail="No historical data file found")
    if body.get("clear_existing"):
        await db.schedule_cells.delete_many({})
    therapists = await db.therapists.find({}, {"_id": 0}).to_list(100)
    t_by_name = {t["name"]: t["id"] for t in therapists}
    DAYS_MAP = {"Sunday":0, "Monday":1, "Tuesday":2, "Wednesday":3, "Thursday":4}
    TIMES = ["8:00 AM - 9:00 AM","9:00 AM - 10:00 AM","10:00 AM - 11:00 AM",
             "11:00 AM - 12:00 PM","12:00 PM - 1:00 PM","1:00 PM - 2:00 PM",
             "2:00 PM - 3:00 PM","3:00 PM - 4:00 PM","4:00 PM - 5:00 PM",
             "5:00 PM - 6:00 PM"]
    inserted = 0
    weeks_loaded = 0
    for week_label, therapists_data in data.items():
        # parse week label like "26 Apr- 30 Apr" → use a fake ISO date for storage
        week_start_iso = f"hist:{week_label}"
        weeks_loaded += 1
        for entry in therapists_data:
            tname = entry.get("n")
            t_id = t_by_name.get(tname)
            if not t_id:
                continue
            for day_label, slots in entry.get("s", []):
                day_idx = DAYS_MAP.get(day_label)
                if day_idx is None:
                    continue
                for slot_idx, raw in enumerate(slots):
                    if not raw or not str(raw).strip():
                        continue
                    txt = str(raw).strip()
                    # parse service code
                    service = "SS"
                    child = None
                    note = None
                    custom = None
                    upper = txt.upper()
                    if upper.startswith("HS"): service = "HS"
                    elif upper.startswith("SS"): service = "SS"
                    elif upper.startswith("OS"): service = "OS"
                    elif "AVC" in upper: service = "AVC"
                    elif "SUPERVISION" in upper: service = "SUPERVISION"
                    elif "OBSERVATION" in upper: service = "OBSERVATION"
                    elif "MEETING" in upper: service = "MEETING"
                    elif "LEAVE" in upper: service = "LEAVE"
                    elif "BREAK" in upper: service = "BREAK"
                    # extract child name after | or W/
                    if "|" in txt:
                        child = txt.split("|", 1)[1].strip()
                    elif "W/" in txt:
                        child = txt.split("W/", 1)[1].strip()
                    elif "with" in txt.lower():
                        child = txt.lower().split("with", 1)[1].strip()
                    if child and "(" in child:
                        custom = child[child.find("(")+1:child.find(")")]
                        child = child[:child.find("(")].strip()
                    if slot_idx >= len(TIMES):
                        continue
                    if service in ("LEAVE", "BREAK", "AVC"):
                        note = txt
                    await db.schedule_cells.insert_one({
                        "id": str(uuid.uuid4()),
                        "therapist_id": t_id, "day": day_idx,
                        "time_slot": TIMES[slot_idx],
                        "service_code": service, "child_name": child,
                        "note": note, "custom_time": custom,
                        "state": "normal", "color": None, "duration": 1,
                        "week_start": week_start_iso, "created_at": now_iso(),
                    })
                    inserted += 1
    return {"weeks_loaded": weeks_loaded, "cells_inserted": inserted}

@api.post("/schedule/duplicate-week")
async def duplicate_week(body: dict, _=Depends(admin_only)):
    """Copy all cells from source_week to target_week. body: {source_week, target_week, clear_target?}"""
    source = body.get("source_week"); target = body.get("target_week")
    if not source or not target:
        raise HTTPException(status_code=400, detail="source_week and target_week required")
    if body.get("clear_target"):
        await db.schedule_cells.delete_many({"week_start": target})
    cells = await db.schedule_cells.find({"week_start": source}, {"_id": 0}).to_list(5000)
    inserted = 0
    for c in cells:
        new_c = {**c, "id": str(uuid.uuid4()), "week_start": target,
                 "state": "normal", "created_at": now_iso()}
        await db.schedule_cells.insert_one(new_c)
        inserted += 1
    return {"copied": inserted}

@api.post("/import/schedule-excel")
async def import_schedule_excel(file: UploadFile = File(...),
                                 week_start: str = Form(...),
                                 clear_existing: Optional[str] = Form(None),
                                 _=Depends(admin_only)):
    """Parse a Therapists' Schedule .xlsx file and create cells for the given week_start.
    Expected layout: each sheet/section has a therapist name with rows for Sunday→Thursday
    and 10 time-slot columns. Cell text format: 'SS | Child', 'HS | Child', 'Meeting w/ X', 'AVC', 'Leave', etc.
    """
    import openpyxl, io
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    therapists = await db.therapists.find({}, {"_id": 0, "pin_hash": 0}).to_list(100)
    t_by_name = {t["name"]: t["id"] for t in therapists}
    # Also accept short names
    for t in therapists:
        short = t["name"].replace("Ms. ", "").strip()
        t_by_name[short] = t["id"]
        t_by_name[short.lower()] = t["id"]

    DAYS_MAP = {"sunday": 0, "monday": 1, "tuesday": 2, "wednesday": 3, "thursday": 4}
    TIMES = ["8:00 AM - 9:00 AM","9:00 AM - 10:00 AM","10:00 AM - 11:00 AM",
             "11:00 AM - 12:00 PM","12:00 PM - 1:00 PM","1:00 PM - 2:00 PM",
             "2:00 PM - 3:00 PM","3:00 PM - 4:00 PM","4:00 PM - 5:00 PM",
             "5:00 PM - 6:00 PM"]

    if clear_existing == "true":
        await db.schedule_cells.delete_many({"week_start": week_start})

    inserted = 0
    skipped_unknown_therapist = []

    def parse_cell(txt: str):
        """Returns (service_code, child_name, custom_time, note) for a cell text."""
        if not txt or not str(txt).strip():
            return None
        txt = str(txt).strip()
        upper = txt.upper()
        custom = None
        note = None
        child = None
        service = "SS"
        if "AVC" in upper: service = "AVC"; note = txt
        elif "LEAVE" in upper: service = "LEAVE"; note = txt
        elif "BREAK" in upper: service = "BREAK"; note = txt
        elif "SUPERVISION" in upper: service = "SUPERVISION"
        elif "OBSERVATION" in upper: service = "OBSERVATION"
        elif "MEETING" in upper: service = "MEETING"
        elif upper.startswith("HS"): service = "HS"
        elif upper.startswith("OS"): service = "OS"
        elif upper.startswith("SS"): service = "SS"
        # Extract child after | or W/ or with
        if "|" in txt:
            child = txt.split("|", 1)[1].strip()
        elif "W/" in upper:
            idx = upper.index("W/")
            child = txt[idx+2:].strip()
        elif " with " in txt.lower():
            child = txt.lower().split(" with ", 1)[1].strip().title()
        # Extract custom time inside ( )
        if child and "(" in child:
            m_open = child.find("(")
            m_close = child.find(")", m_open)
            if m_close > m_open:
                custom = child[m_open+1:m_close].strip()
                child = child[:m_open].strip()
        return service, child, custom, note

    # Iterate all sheets
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        i = 0
        while i < len(rows):
            row = rows[i] or ()
            cells = [str(c).strip() if c is not None else "" for c in row]
            # Detect header row: contains "Therapist's Name" and "Days"
            joined = " ".join(c.lower() for c in cells)
            if "therapist" in joined and "days" in joined and "8:00" in joined:
                # The next 5 rows should be: therapist_name | day | 10 slots
                current_t_id = None
                for k in range(1, 6):
                    if i + k >= len(rows):
                        break
                    sub = rows[i + k] or ()
                    sub_cells = [str(c).strip() if c is not None else "" for c in sub]
                    if len(sub_cells) < 4:
                        continue
                    # Therapist name typically in col B (idx 1) on first sub-row
                    name_candidate = sub_cells[1] if len(sub_cells) > 1 else ""
                    if name_candidate and current_t_id is None:
                        if name_candidate in t_by_name:
                            current_t_id = t_by_name[name_candidate]
                        else:
                            for key, tid in t_by_name.items():
                                if key.lower() == name_candidate.lower():
                                    current_t_id = tid; break
                    if current_t_id is None:
                        # Skip if can't find therapist; but track for skipped
                        if name_candidate and name_candidate not in skipped_unknown_therapist:
                            skipped_unknown_therapist.append(name_candidate)
                        continue
                    # Day in col C (idx 2)
                    day_label = sub_cells[2].lower() if len(sub_cells) > 2 else ""
                    day_idx = DAYS_MAP.get(day_label)
                    if day_idx is None:
                        continue
                    # Time slots in cols D-M (idx 3-12)
                    for slot_idx in range(10):
                        col_idx = 3 + slot_idx
                        if col_idx >= len(sub_cells):
                            break
                        val = sub_cells[col_idx]
                        parsed = parse_cell(val)
                        if not parsed:
                            continue
                        service, child, custom, note = parsed
                        await db.schedule_cells.insert_one({
                            "id": str(uuid.uuid4()),
                            "therapist_id": current_t_id,
                            "day": day_idx, "time_slot": TIMES[slot_idx],
                            "service_code": service, "child_name": child,
                            "note": note, "custom_time": custom,
                            "state": "normal", "color": None, "duration": 1,
                            "week_start": week_start, "created_at": now_iso(),
                        })
                        inserted += 1
                i += 6  # skip past header + 5 data rows
                continue
            i += 1

    return {"cells_inserted": inserted, "week_start": week_start, "skipped_therapists": skipped_unknown_therapist[:20]}

@api.get("/")
async def root():
    return {"message": "Boost Growth Portal API", "status": "ok"}

app.include_router(api)
app.add_middleware(CORSMiddleware, allow_credentials=True, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ------------------- Seed Data (FROM BASE44 SOURCE) -------------------
THERAPIST_SEED = [
    {"name": "Ms. Maha", "color": "#7A8A6A", "email": "maha@boostgrowthsa.com"},
    {"name": "Ms. Fahda", "color": "#D4A64A", "email": "fahda@boostgrowthsa.com"},
    {"name": "Ms. Razan", "color": "#8FA481", "email": "razan@boostgrowthsa.com"},
    {"name": "Ms. Manal", "color": "#A4BCCB", "email": "manal@boostgrowthsa.com"},
    {"name": "Ms. Hajer", "color": "#C97B5C", "email": "hajer@boostgrowthsa.com"},
    {"name": "Ms. Rahaf", "color": "#9B7BAB", "email": "rahaf@boostgrowthsa.com"},
    {"name": "Ms. Shatha", "color": "#5C8B7E", "email": "shatha@boostgrowthsa.com"},
    {"name": "Ms. Alhanouf", "color": "#B89968", "email": "alhanouf@boostgrowthsa.com"},
    {"name": "Ms. Waad", "color": "#7B96B5", "email": "waad@boostgrowthsa.com"},
    {"name": "Ms. Bodoor", "color": "#A8745E", "email": "bodoor@boostgrowthsa.com"},
    {"name": "Ms. Fatimah", "color": "#6B9080", "email": "fatimah@boostgrowthsa.com"},
    {"name": "Ms. Shrooq", "color": "#D49A60", "email": "shrooq@boostgrowthsa.com"},
    {"name": "Ms. Abeer", "color": "#8B7BA8", "email": "abeer@boostgrowthsa.com"},
    {"name": "Ms. Najla", "color": "#7BA890", "email": "najla@boostgrowthsa.com"},
    {"name": "Ms. Walaa", "color": "#C28E6A", "email": "walaa@boostgrowthsa.com"},
    {"name": "Ms. Asma", "color": "#6A7F9B", "email": "asma@boostgrowthsa.com"},
    {"name": "Ms. Jenan", "color": "#A38B5F", "email": "jenan@boostgrowthsa.com"},
]

CLIENT_SEED = [
    {"file_no":"009","name":"Saleh Ahusainy","main":"Ms. Waad","co":["Ms. Manal","Ms. Fahda"],"pkg":24,"sup":"Ms. Fahda","color":"#FFE599","locs":[{"service":"SS","address":"Alnakeel - Home Sweet Home"},{"service":"HS","address":"Alnakheel - 1st floor, apartment #7"},{"service":"HS","address":"Grandmother house"}]},
    {"file_no":"011","name":"Fahad Alyahya","main":"Ms. Alhanouf","co":["Ms. Fahda"],"pkg":24,"sup":"Ms. Fahda","color":"#A2C4C9","locs":[{"service":"HS","address":"Alyasmin - house no 3075"},{"service":"SS","address":"Talat School"}]},
    {"file_no":"018","name":"Layan AlSaud","main":"Ms. Jenan","co":[],"pkg":24,"sup":"Ms. Jenan","color":"#C9DAF8","locs":[{"service":"ABA","address":"Alaqiq"}]},
    {"file_no":"023","name":"Yahya Alqahtani","main":"Ms. Hajer","co":["Ms. Manal"],"pkg":24,"sup":"Ms. Fahda","color":"#D5A6BD","locs":[{"service":"HS","address":"Alaarid"}]},
    {"file_no":"024","name":"Abdulaziz Alrasheed","main":"Ms. Shatha","co":["Ms. Manal","Ms. Hajer"],"pkg":24,"sup":"Ms. Fahda","color":"#E6B8AF","locs":[{"service":"HS","address":"Alnada - Building #26, 3rd floor, apartment #23"}]},
    {"file_no":"027","name":"Mohmmed Alaqel","main":"Ms. Rahaf","co":["Ms. Fahda"],"pkg":24,"sup":"Ms. Fahda","color":"#FFF2CC","locs":[{"service":"HS","address":"AlMalqa - 331"}]},
    {"file_no":"030","name":"Husam Alturaigy","main":"Ms. Manal","co":["Ms. Shatha"],"pkg":24,"sup":"Ms. Fahda","color":"#B4A7D6","locs":[{"service":"SS","address":"Whales of the future daycare"},{"service":"HS","address":"Alwaha - Home #4B"}]},
    {"file_no":"034","name":"Aljouhrah Alduailij","main":"Ms. Asma","co":["Ms. Fahda"],"pkg":24,"sup":"Ms. Fahda","color":"#D9EAD3","locs":[{"service":"SS","address":"Alnakheel - Talat School"}]},
    {"file_no":"035","name":"Saad Alghamdi","main":"Ms. Shatha","co":["Ms. Hajer","Ms. Fatimah"],"pkg":24,"sup":"Ms. Maha","color":"#B6D7A8","locs":[{"service":"HS","address":"Al Aqiq - House in the corner"},{"service":"SS","address":"Al Motaqdimah Schools"}]},
    {"file_no":"037","name":"Suzan Alsultan","main":"Ms. Asma","co":[],"pkg":24,"sup":"Ms. Maha","color":"#FCE5CD","locs":[{"service":"HS","address":"King Fahad - Villa 1308"}]},
    {"file_no":"038","name":"Salman Alrasheed","main":"Ms. Manal","co":["Ms. Fahda"],"pkg":24,"sup":"Ms. Maha","color":"#F4CCCC","locs":[{"service":"SS","address":"Summer Camp - Stars of Knowledge School"},{"service":"HS","address":"Alnada - Building #26, 3rd floor, apartment #23"}]},
    {"file_no":"040","name":"Abdulaziz AlAbdulwahab","main":"Ms. Fatimah","co":["Ms. Fahda","Ms. Hajer"],"pkg":24,"sup":"Ms. Maha","color":"#6FA8DC","locs":[{"service":"HS","address":"Alraed - house no 8188"}]},
    {"file_no":"041","name":"Ameerah Alshehri","main":"Ms. Fahda","co":["Ms. Fatimah"],"pkg":24,"sup":"Ms. Maha","color":"#EA9999","locs":[{"service":"HS","address":"Roshen - Villa 277"}]},
    {"file_no":"042","name":"Sultan Aldamer","main":"Ms. Shrooq","co":["Ms. Rahaf"],"pkg":24,"sup":"Ms. Maha","color":"#FFE599","locs":[{"service":"SS","address":"Bright Mind School"},{"service":"HS","address":"Alhada - No house number"}]},
    {"file_no":"047","name":"Alwaleed Alotaibi","main":"Ms. Hajer","co":["Ms. Alhanouf"],"pkg":24,"sup":"Ms. Maha","color":"#B4A7D6","locs":[{"service":"HS","address":"Alqairawan - house no 10"},{"service":"SS","address":"Al Motaqdimah Schools"}]},
    {"file_no":"052","name":"Sulaiman Alkhurashi","main":"Ms. Rahaf","co":["Ms. Maha"],"pkg":24,"sup":"Ms. Maha","color":"#F9CB9C","locs":[{"service":"HS","address":"Alsulaimanyah - house no 24"}]},
    {"file_no":"054","name":"Omar Alkhurashi","main":"Ms. Manal","co":["Ms. Maha"],"pkg":24,"sup":"Ms. Maha","color":"#D0E0E3","locs":[{"service":"HS","address":"Alsulaimanyah - house no 24"}]},
    {"file_no":"060","name":"Mohammed Albedayea","main":"Ms. Bodoor","co":["Ms. Shatha"],"pkg":24,"sup":"Ms. Maha","color":"#D9EAD3","locs":[{"service":"HS","address":"Alyasmin - Home no 14"},{"service":"SS","address":"Yas School"}]},
    {"file_no":"061","name":"Ibrahim Alnasir","main":"Ms. Rahaf","co":["Ms. Fahda"],"pkg":24,"sup":"Ms. Fahda","color":"#D9D2E9","locs":[{"service":"HS","address":"Alyasmin - Home no 39"},{"service":"SS","address":"Alnakheel - Talat School"}]},
    {"file_no":"062","name":"Lulu Almutair","main":"Ms. Razan","co":["Ms. Fahda"],"pkg":24,"sup":"Ms. Fahda","color":"#D5A6BD","locs":[{"service":"HS","address":"Almuroj - Home no 4"},{"service":"SS","address":"Alnakheel - Talat School"}]},
    {"file_no":"063","name":"Amani Ghaith","main":"Ms. Maha","co":[],"pkg":24,"sup":"Ms. Maha","color":"#FFF2CC","locs":[{"service":"HS","address":"Alnakheel"}]},
    {"file_no":"065","name":"Aser Alharbi","main":"Ms. Najla","co":["Ms. Maha"],"pkg":24,"sup":"Ms. Maha","color":"#F4CCCC","locs":[{"service":"HS","address":"Al Izdihar - First floor - House no 15"}]},
    {"file_no":"068","name":"Abdulrahman Alshawi","main":"Ms. Razan","co":["Ms. Fahda"],"pkg":24,"sup":"Ms. Fahda","color":"#C9DAF8","locs":[{"service":"HS","address":"AR Rayan - Home no 32"},{"service":"SS","address":"Kindergarten of KSU"}]},
    {"file_no":"070","name":"Abdulelah Almuhana","main":"Ms. Abeer","co":["Ms. Maha"],"pkg":24,"sup":"Ms. Maha","color":"#CFE2F3","locs":[{"service":"SS","address":"Manarat Ar Riyadh"}]},
    {"file_no":"072","name":"Khalid Bin Shuael","main":"Ms. Shatha","co":["Ms. Fahda"],"pkg":24,"sup":"Ms. Fahda","color":"#EAD1DC","locs":[{"service":"HS","address":"AlMursalat"}]},
]

# ------------------- Intake Seed (from Waiting_List_v4.xlsx) -------------------
INTAKE_SEED = [
    # Pre-Intake
    {"intake_type":"pre","child_name":"Reema Idrees","service":"HS","phone":"546272994","district":"Iraqi","age":"2021","time_pref":"Morning","diagnosis":"PWS"},
    {"intake_type":"pre","child_name":"Abdulaziz Alrajab","service":"HS","phone":"500252211","district":"Al Malqa","age":"2023","time_pref":"Any","diagnosis":"NA","notes":"Online CONCL"},
    {"intake_type":"pre","child_name":"Mansour","service":"HS","phone":"507247881","district":"Alyasmeen","age":"2022","time_pref":"Any","diagnosis":"Speech delay"},
    {"intake_type":"pre","child_name":"Leen","service":"SS","phone":"503225528","district":"Al Raed","age":"2010","time_pref":"Morning","diagnosis":"NA","notes":"3 hours at school"},
    {"intake_type":"pre","child_name":"Ebrahim Alnami","service":"SS","phone":"564443542","district":"Alsulimania","age":"2022","time_pref":"Morning","diagnosis":"Premature - 29 weeks"},
    {"intake_type":"pre","child_name":"Naif Alblawi","service":"HS","phone":"535544260","district":"Qurtubah","age":"2020","time_pref":"Evening","diagnosis":"ADHD"},
    {"intake_type":"pre","child_name":"Saad Alajaji","service":"HS","phone":"555955342","district":"AL-Suwaidi","age":"2021","time_pref":"Evening","diagnosis":"NA"},
    {"intake_type":"pre","child_name":"Reema Alotaibi","service":"HS","phone":"503553339","district":"AlArid","time_pref":"Evening","diagnosis":"Speech delay"},
    {"intake_type":"pre","child_name":"Waseem Aljohani","service":"HS / SS","phone":"594744884","district":"Alnarjis","age":"2019","diagnosis":"ADHD","notes":"DR.Turki"},
    {"intake_type":"pre","child_name":"Faisal Alzghaibi","service":"HS","phone":"966507479800","district":"Alyasmeen","diagnosis":"NA","notes":"Azraq"},
    {"intake_type":"pre","child_name":"Sultan Abalkhail","service":"HS","phone":"558811313","district":"Al-Mursalat","age":"2019","diagnosis":"NA"},
    {"intake_type":"pre","child_name":"Sultan Bandar","service":"HS","phone":"555579702","district":"Alyasmeen","age":"2019","time_pref":"Any","diagnosis":"Speech delay - ADHD"},
    # Post-Intake
    {"intake_type":"post","child_name":"Mohammed alnoweser","service":"HS","district":"King Fahad","age":"3 year","language":"English"},
    {"intake_type":"post","child_name":"Mohammed Alofi","service":"HS","phone":"554505400","district":"AlAridh","age":"6","language":"English / Arabic"},
    {"intake_type":"post","child_name":"Rakan Alaqel","service":"HS","phone":"538154083","district":"Alnarjis","age":"2019","language":"Arabic"},
    {"intake_type":"post","child_name":"Nawaf Alshweeb","service":"HS","district":"Um Alhamam","age":"5.5","language":"ASD"},
    {"intake_type":"post","child_name":"Abdulkareem Kaki","service":"HS","language":"Arabic"},
    {"intake_type":"post","child_name":"Abdulaziz Alzahrani","service":"HS","phone":"555341092","district":"Almalqa","age":"4"},
    {"intake_type":"post","child_name":"Yazeed Bu sheet","service":"SS","phone":"555009662","district":"Hitter","diagnosis":"Autism"},
    {"intake_type":"post","child_name":"Omar ALImazrou","service":"HS","phone":"534888855","district":"AlArid","age":"2023","diagnosis":"Autism"},
    {"intake_type":"post","child_name":"Fahad Suliman","service":"HS","phone":"966500566235","district":"Al-Sahafa","age":"2019","diagnosis":"ADD"},
    {"intake_type":"post","child_name":"Naif Alwhibi","service":"SS / HS","phone":"506128118","district":"Ar Rabi","age":"2020","diagnosis":"ASD"},
    {"intake_type":"post","child_name":"Ahmad Alshalfan","service":"SS / HS","phone":"505287407","district":"Almalqa","age":"2020","diagnosis":"ADHD and GDD"},
    {"intake_type":"post","child_name":"Abdulelah Almuhana","service":"HS","phone":"966 56 554 4999","age":"2021"},
    {"intake_type":"post","child_name":"Leena Alshahrani","service":"HS","phone":"530511175"},
]

# ------------------- Directory Seed (Internal Team) -------------------
DIRECTORY_SEED = [
    {"name":"Genan Almuhaisen","role":"Direct Manager","phone":"","email":"genan@boostgrowthsa.com"},
    {"name":"Boost Growth (Main)","role":"Coordinator / General Inquiries","phone":"","email":"hello@boostgrowthsa.com"},
    {"name":"Ms. Walaa","role":"Operations","phone":"","email":"walaa@boostgrowthsa.com"},
    {"name":"Ms. Maha","role":"Supervisor","phone":"","email":"maha@boostgrowthsa.com"},
    {"name":"Ms. Fahdah","role":"Supervisor","phone":"","email":"fahda@boostgrowthsa.com"},
]

# ------------------- Resources Seed -------------------
RESOURCES_SEED = [
    {"title":"Therapist Drive","description":"Session materials · forms · training","url":"https://drive.google.com/drive/folders/1iMDwfucwzsEIl9WxwhJi_h6tg2vVtAFr","visibility":"therapist","icon":"Folders","bg":"#E5EBE1","color":"#3D4F35","sort_order":10},
    {"title":"Therapist Training Hub","description":"Protocols · lesson plans","url":"https://drive.google.com/drive/folders/1iMDwfucwzsEIl9WxwhJi_h6tg2vVtAFr","visibility":"therapist","icon":"Notebook","bg":"#EAF0F3","color":"#375568","sort_order":20},
    {"title":"Client Files","description":"Per-client folders","url":"https://drive.google.com/drive/folders/1iMDwfucwzsEIl9WxwhJi_h6tg2vVtAFr","visibility":"admin","icon":"Folders","bg":"#FAF0D1","color":"#6B5218","sort_order":30},
    {"title":"HR Files","description":"Employees · Contracts","url":"https://drive.google.com/drive/folders/1jWRO97gDHK_TfmZhTqCqm0SdBc6_b5bE","visibility":"admin","icon":"Files","bg":"#F1ECF7","color":"#4E3F70","sort_order":40},
    {"title":"Company Policies","description":"Internal policies & SOPs","url":"https://drive.google.com/drive/folders/11VQQ-o1QoDQV-ktygB1tlnRmqCs3mxAb","visibility":"all","icon":"Notebook","bg":"#F4E7D8","color":"#8B6918","sort_order":50},
]


@app.on_event("startup")
async def startup():
    await db.users.create_index("email", unique=True)
    await db.therapists.create_index("id", unique=True)
    await db.schedule_cells.create_index([("week_start", 1), ("therapist_id", 1)])
    await db.notifications.create_index("user_id")
    await db.sessions.create_index([("client_id", 1), ("session_date", -1)])

    admin_email = os.environ["ADMIN_EMAIL"].lower()
    admin_password = os.environ["ADMIN_PASSWORD"]
    admin_name = os.environ.get("ADMIN_NAME", "Admin")
    existing = await db.users.find_one({"email": admin_email})
    if not existing:
        await db.users.insert_one({"id": str(uuid.uuid4()), "email": admin_email,
                                   "password_hash": hash_password(admin_password),
                                   "name": admin_name, "role": "admin", "created_at": now_iso()})
        logger.info(f"Admin seeded: {admin_email}")
    elif not verify_password(admin_password, existing["password_hash"]):
        await db.users.update_one({"email": admin_email}, {"$set": {"password_hash": hash_password(admin_password)}})

    # Seed therapists if missing or count mismatch
    th_count = await db.therapists.count_documents({})
    if th_count != len(THERAPIST_SEED):
        await db.therapists.delete_many({})
        for s in THERAPIST_SEED:
            await db.therapists.insert_one({
                "id": str(uuid.uuid4()), "name": s["name"], "color": s["color"],
                "email": s.get("email"), "phone": None,
                "pin_hash": hash_password("0000"),
                "created_at": now_iso(),
            })
        logger.info(f"Seeded {len(THERAPIST_SEED)} therapists with PIN=0000")

    # Load persisted email settings from db.settings into env
    settings_doc = await db.settings.find_one({"key": "email"}, {"_id": 0})
    if settings_doc:
        if settings_doc.get("resend_api_key"):
            os.environ["RESEND_API_KEY"] = settings_doc["resend_api_key"]
        if settings_doc.get("from_email"):
            os.environ["EMAIL_FROM"] = settings_doc["from_email"]

    # Re-seed clients with full info (force re-seed if seed data changed — track version)
    seed_meta = await db.meta.find_one({"key": "client_seed_version"})
    CURRENT_SEED_VERSION = 6  # bump when CLIENT_SEED changes
    cl_count = await db.clients.count_documents({})
    if cl_count != len(CLIENT_SEED) or not seed_meta or seed_meta.get("version") != CURRENT_SEED_VERSION:
        await db.clients.delete_many({})
        therapists_map = {t["name"]: t["id"] async for t in db.therapists.find({}, {"_id": 0, "name": 1, "id": 1})}
        for c in CLIENT_SEED:
            await db.clients.insert_one({
                "id": str(uuid.uuid4()),
                "file_no": c["file_no"], "name": c["name"],
                "package_hours": c["pkg"], "supervisor": c["sup"],
                "main_therapist_id": therapists_map.get(c["main"]),
                "co_therapist_ids": [therapists_map[n] for n in c["co"] if n in therapists_map],
                "color": c["color"], "locations": c["locs"],
                "parent_name": None, "parent_phone": None, "age": None,
                "notes": None, "created_at": now_iso(),
            })
        await db.meta.update_one({"key": "client_seed_version"},
                                 {"$set": {"version": CURRENT_SEED_VERSION, "updated_at": now_iso()}},
                                 upsert=True)
        logger.info(f"Seeded {len(CLIENT_SEED)} clients (v{CURRENT_SEED_VERSION}) with accurate multi-service info")

    # Seed Intake (only if empty — admin may manage manually)
    if await db.intake.count_documents({}) == 0:
        for item in INTAKE_SEED:
            await db.intake.insert_one({
                "id": str(uuid.uuid4()),
                "status": "new",
                "priority": False,
                "created_at": now_iso(),
                **item,
            })
        logger.info(f"Seeded {len(INTAKE_SEED)} intake records from waiting list")

    # Seed Directory (only if empty)
    if await db.directory.count_documents({}) == 0:
        for item in DIRECTORY_SEED:
            await db.directory.insert_one({
                "id": str(uuid.uuid4()), **item, "created_at": now_iso(),
            })
        logger.info(f"Seeded {len(DIRECTORY_SEED)} directory contacts")

    # Seed Resources (only if empty)
    if await db.resources.count_documents({}) == 0:
        for item in RESOURCES_SEED:
            await db.resources.insert_one({
                "id": str(uuid.uuid4()),
                "category": "drive",
                **item,
                "created_at": now_iso(),
            })
        logger.info(f"Seeded {len(RESOURCES_SEED)} resources")

    # Seed Leaves (only if empty) — from leaves_seed.json (parsed Vacation 2026)
    if await db.leaves.count_documents({}) == 0:
        seed_path = ROOT_DIR / "leaves_seed.json"
        if seed_path.exists():
            import json
            seed = json.loads(seed_path.read_text())
            t_by_name = {t["name"]: t["id"] async for t in db.therapists.find({}, {"_id": 0, "name": 1, "id": 1})}
            inserted = 0
            for item in seed:
                tid = t_by_name.get(item.get("therapist_name"))
                if not tid:
                    continue
                await db.leaves.insert_one({
                    "id": str(uuid.uuid4()),
                    "therapist_id": tid,
                    "start_date": item.get("start_date") or "",
                    "end_date": item.get("end_date") or "",
                    "days": item.get("days") or 0,
                    "leave_type": item.get("leave_type") or "Annual",
                    "status": item.get("status") or "done",
                    "notes": item.get("notes"),
                    "created_at": now_iso(),
                })
                inserted += 1
            logger.info(f"Seeded {inserted} leaves from Vacation 2026")

@app.on_event("shutdown")
async def shutdown():
    client.close()
